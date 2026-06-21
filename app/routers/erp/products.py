"""Products + variants endpoints. Thin: validate input, delegate to ProductsService."""

from __future__ import annotations

import io

from fastapi import APIRouter, Depends, File, Query, UploadFile
from fastapi.responses import StreamingResponse

from app.models.erp_schemas import (
    ImportResult,
    ProductCreate,
    ProductImageAnalysis,
    ProductResponse,
    ProductUpdate,
    VariantCreate,
    VariantResponse,
)
from app.services.erp.context import ERPContext, get_erp_context, log_activity
from app.services.erp.exceptions import ValidationError
from app.services.erp.products import ProductsService
from app.services.images import optimize_image
from app.services.storage import upload_product_image
from app.services.vision import analyze_product_image

router = APIRouter()
service = ProductsService()

_IMPORT_COLUMNS = ["name", "category", "sku", "barcode", "cost_price", "price", "unit", "low_stock_threshold"]


@router.get("", response_model=list[ProductResponse])
async def list_products(
    ctx: ERPContext = Depends(get_erp_context),
    category: str | None = None,
    search: str | None = None,
    limit: int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0),
):
    return await service.list(ctx, category=category, search=search, limit=limit, offset=offset)


@router.post("", response_model=ProductResponse)
async def create_product(body: ProductCreate, ctx: ERPContext = Depends(get_erp_context)):
    return await service.create(ctx, body.model_dump(exclude_none=True))


@router.post("/analyze-image", response_model=ProductImageAnalysis)
async def analyze_image(
    file: UploadFile = File(...), ctx: ERPContext = Depends(get_erp_context)
):
    """Optimiza la imagen (WebP cuadrado), la sube a Storage y la analiza con Gemini.

    NO crea el producto: devuelve la URL pública + nombre/descripción/tags sugeridos para
    que el front los edite y luego guarde con POST /erp/products. Si Gemini falla o no está
    configurado, igual devuelve la `image_url` con `ai_ok=false`.
    """
    optimized = optimize_image(await file.read())
    image_url = upload_product_image(ctx.tenant_id, optimized)
    analysis = analyze_product_image(optimized, "image/webp")
    log_activity(ctx, action="product.image_analyzed", module="inventory",
                 detail={"ai_ok": analysis["ai_ok"]})
    return ProductImageAnalysis(image_url=image_url, **analysis)


@router.get("/barcode/{code}", response_model=ProductResponse)
async def get_by_barcode(code: str, ctx: ERPContext = Depends(get_erp_context)):
    return await service.get_by_barcode(ctx, code)


@router.get("/import/template")
async def import_template(_ctx: ERPContext = Depends(get_erp_context)):
    """Download an .xlsx template with the expected columns."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"
    ws.append(_IMPORT_COLUMNS)
    ws.append(["Heineken 1L", "Cervezas", "HEI-1L", "7791234567890", 18.0, 25.0, "unidad", 6])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="plantilla_productos.xlsx"'},
    )


@router.post("/import", response_model=ImportResult)
async def import_products(
    file: UploadFile = File(...), ctx: ERPContext = Depends(get_erp_context)
):
    """Import products row by row. Bad rows are skipped and reported, good rows are created."""
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(await file.read()), read_only=True, data_only=True)
    except Exception as exc:
        raise ValidationError("No se pudo leer el archivo Excel", reason=str(exc))

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return ImportResult(imported=0, errors=[])

    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    imported, errors = 0, []
    for i, raw in enumerate(rows[1:], start=2):
        record = {header[j]: raw[j] for j in range(min(len(header), len(raw)))}
        try:
            if not record.get("name"):
                raise ValueError("nombre vacío")
            payload = ProductCreate(
                name=str(record["name"]).strip(),
                category=(str(record["category"]).strip() if record.get("category") else None),
                sku=(str(record["sku"]).strip() if record.get("sku") else None),
                barcode=(str(record["barcode"]).strip() if record.get("barcode") else None),
                cost_price=float(record.get("cost_price") or 0),
                price=float(record.get("price") or 0),
                unit=(str(record["unit"]).strip() if record.get("unit") else "unidad"),
                low_stock_threshold=int(record.get("low_stock_threshold") or 5),
            )
            await service.create(ctx, payload.model_dump(exclude_none=True))
            imported += 1
        except Exception as exc:  # noqa: BLE001 — per-row, reported not raised
            errors.append({"row": i, "reason": str(exc)})

    log_activity(ctx, action="products.imported", module="inventory",
                 detail={"imported": imported, "error_count": len(errors)})
    return ImportResult(imported=imported, errors=errors)


@router.get("/{product_id}", response_model=ProductResponse)
async def get_product(product_id: str, ctx: ERPContext = Depends(get_erp_context)):
    return await service.get(ctx, product_id)


@router.put("/{product_id}", response_model=ProductResponse)
async def update_product(
    product_id: str, body: ProductUpdate, ctx: ERPContext = Depends(get_erp_context)
):
    return await service.update(ctx, product_id, body.model_dump(exclude_unset=True))


@router.delete("/{product_id}")
async def delete_product(product_id: str, ctx: ERPContext = Depends(get_erp_context)):
    return await service.soft_delete(ctx, product_id)


@router.post("/{product_id}/variants", response_model=VariantResponse)
async def add_variant(
    product_id: str, body: VariantCreate, ctx: ERPContext = Depends(get_erp_context)
):
    return await service.add_variant(ctx, product_id, body.model_dump(exclude_none=True))


@router.put("/{product_id}/variants/{variant_id}", response_model=VariantResponse)
async def update_variant(
    product_id: str, variant_id: str, body: VariantCreate,
    ctx: ERPContext = Depends(get_erp_context),
):
    return await service.update_variant(ctx, product_id, variant_id, body.model_dump(exclude_unset=True))
